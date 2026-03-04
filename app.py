import streamlit as st
import pandas as pd
import io
import os
import openpyxl
from openpyxl.drawing.image import Image as xlImage
from openpyxl.styles import Alignment, Border, Side
from copy import copy

st.set_page_config(page_title="Esclerometria · Tecomat", layout="wide", page_icon="🏗️")

st.markdown("""
<style>
    .main { background-color: #0f1923; }
    h1 { color: #f0a500; }
    .stButton>button { background-color: #f0a500; color: black; font-weight: bold; width: 100%; }
    .stDownloadButton>button { background-color: #00c9a7; color: black; font-weight: bold; width: 100%; font-size: 1.1rem; padding: 0.75rem; }
    .stTabs [data-baseweb="tab"] { font-size: 1rem; font-weight: bold; }
</style>
""", unsafe_allow_html=True)

st.title("🏗️ Esclerometria · Tecomat")

# Memória do app
if 'amostras' not in st.session_state:
    st.session_state.amostras = []
if 'coef_bigorna' not in st.session_state:
    st.session_state.coef_bigorna = 1.0
if 'media_bigorna' not in st.session_state:
    st.session_state.media_bigorna = 0.0

aba1, aba2, aba3 = st.tabs(["📋 Cabeçalho e Aparelho", "🔨 Dados de Campo", "📄 Exportar Relatório"])

# ══════════════════════════════════════════
# ABA 1 — CABEÇALHO
# ══════════════════════════════════════════
with aba1:
    st.subheader("Dados da Obra e Cliente")
    col1, col2 = st.columns(2)
    rlt          = col1.text_input("Nº do RLT", placeholder="Ex: 042")
    data_emissao = col2.text_input("Data de Emissão", placeholder="DD/MM/AAAA")
    cliente      = st.text_input("Cliente")
    obra         = st.text_input("Obra")
    col3, col4   = st.columns(2)
    att          = col3.text_input("Att")
    endereco     = col4.text_input("Endereço")

    st.divider()
    st.subheader("🔩 Verificação do Aparelho (10 Golpes)")

    cols_big = st.columns(10)
    bigorna_vals = []
    for i in range(10):
        v = cols_big[i].text_input(f"G{i+1}", key=f"big_{i}", label_visibility="collapsed", placeholder=str(i+1))
        bigorna_vals.append(v)

    validos_big = []
    for v in bigorna_vals:
        v2 = v.replace(',', '.')
        if v2.replace('.', '', 1).isdigit():
            validos_big.append(float(v2))

    media_big = sum(validos_big) / len(validos_big) if validos_big else 0.0
    coef_big  = 80.0 / media_big if media_big > 0 else 1.0
    st.session_state.media_bigorna  = media_big
    st.session_state.coef_bigorna   = coef_big

    col_m, col_c = st.columns(2)
    col_m.metric("Média", f"{media_big:.2f}")
    col_c.metric("Coef. de Correção", f"{coef_big:.6f}")

    st.divider()
    st.subheader("👤 Responsável Tecomat")
    col5, col6  = st.columns(2)
    resp_nome   = col5.text_input("Nome")
    resp_crea   = col6.text_input("CREA")
    assinatura  = st.file_uploader("Assinatura (PNG/JPG)", type=['png', 'jpg', 'jpeg'])
    notas       = st.text_area("Notas / Observações")

# ══════════════════════════════════════════
# ABA 2 — DADOS DE CAMPO
# ══════════════════════════════════════════
with aba2:
    st.subheader("➕ Nova Amostra")
    col_a, col_b = st.columns([2, 1])
    amostra_nome = col_a.text_input("Identificação", placeholder="Ex: P1 - Térreo")
    posicao      = col_b.selectbox("Posição", ["0° (Horizontal ➔)", "+90° (Para cima ⬆)", "-90° (Para baixo ⬇)"])

    st.write("**Impactos (até 16):**")
    cols_imp1 = st.columns(8)
    cols_imp2 = st.columns(8)
    impactos_vals = []
    for i in range(16):
        col_ref = cols_imp1 if i < 8 else cols_imp2
        idx     = i if i < 8 else i - 8
        v = col_ref[idx].text_input(f"I{i+1}", key=f"imp_{i}", label_visibility="collapsed", placeholder=str(i+1))
        impactos_vals.append(v)

    if st.button("✅ Calcular e Adicionar", type="primary", use_container_width=True):
        if not amostra_nome:
            st.warning("Preencha a identificação da amostra!")
        else:
            validos = []
            for v in impactos_vals:
                v2 = v.replace(',', '.')
                if v2.replace('.', '', 1).isdigit():
                    validos.append(float(v2))

            if not validos:
                st.warning("Insira pelo menos um impacto válido!")
            else:
                ie_bruto = sum(validos) / len(validos)
                lim_inf  = ie_bruto * 0.90
                lim_sup  = ie_bruto * 1.10
                filtrados = [v for v in validos if lim_inf <= v <= lim_sup]

                if len(filtrados) < 5:
                    status     = "Amostra Perdida"
                    ie_filt    = "-"
                    ie_efetivo = "-"
                    fck        = "-"
                    dispersao  = "-"
                else:
                    status     = "Amostra Válida"
                    ie_filt    = sum(filtrados) / len(filtrados)
                    ie_efetivo = ie_filt * st.session_state.coef_bigorna
                    fck        = (0.0106 * (ie_efetivo**2)) + (0.9907 * ie_efetivo) - 13.969
                    if ie_efetivo < 30:   dispersao = "±5,5"
                    elif ie_efetivo < 40: dispersao = "±6,0"
                    elif ie_efetivo < 60: dispersao = "±6,5"
                    else:                 dispersao = "±7,0"

                fmt = lambda v: f"{v:.2f}" if isinstance(v, float) else v
                st.session_state.amostras.append({
                    "Amostra":               amostra_nome,
                    "Posição":               posicao.split(" ")[0],
                    "Lim. Inf.":             fmt(lim_inf),
                    "Lim. Sup.":             fmt(lim_sup),
                    "I.E. Médio":            fmt(ie_filt),
                    "Status":                status,
                    "I.E. Efetivo":          fmt(ie_efetivo),
                    "Resist. Estimada (MPa)":fmt(fck),
                    "Dispersão":             dispersao
                })
                st.success(f"✅ '{amostra_nome}' adicionada!")
                st.rerun()

    st.divider()
    st.subheader("📊 Tabela de Resultados")

    if st.session_state.amostras:
        df = pd.DataFrame(st.session_state.amostras)

        def colorir(row):
            cor = "color: #00c9a7" if row["Status"] == "Amostra Válida" else "color: #e05252"
            return [cor] * len(row)

        st.dataframe(df.style.apply(colorir, axis=1), use_container_width=True)

        validas  = sum(1 for a in st.session_state.amostras if a["Status"] == "Amostra Válida")
        perdidas = len(st.session_state.amostras) - validas
        c1, c2, c3 = st.columns(3)
        c1.metric("Total", len(st.session_state.amostras))
        c2.metric("Válidas", validas)
        c3.metric("Perdidas", perdidas)

        if st.button("🗑️ Limpar toda a tabela"):
            st.session_state.amostras = []
            st.rerun()
    else:
        st.info("Nenhuma amostra inserida ainda.")

# ══════════════════════════════════════════
# ABA 3 — EXPORTAR
# ══════════════════════════════════════════
with aba3:
    st.subheader("📄 Gerar Relatório Excel Oficial")

    # Checklist
    checks = [
        ("Nº RLT preenchido",       bool(rlt)),
        ("Cliente preenchido",       bool(cliente)),
        ("Amostras inseridas",       len(st.session_state.amostras) > 0),
        ("Bigorna verificada",       media_big > 0),
    ]
    for label, ok in checks:
        st.write(f"{'✅' if ok else '⚠️'} {label}")

    st.divider()
    st.info("💡 Após baixar o Excel, abra no **WPS Office** no celular e exporte como PDF — o layout ficará idêntico ao relatório oficial.")

    if not st.session_state.amostras:
        st.warning("Adicione pelo menos uma amostra na aba 'Dados de Campo' antes de exportar.")
    else:
        try:
            caminho_modelo = "Modelo_esclerometria.xlsx"
            wb   = openpyxl.load_workbook(caminho_modelo)
            plan = wb["ESCLEROMETRIA"]

            # RLT
            num_raw = rlt.strip()
            if num_raw.isdigit():
                rlt_oficial = f"RLT.LAU-{int(num_raw):03d}.26-00"
            else:
                rlt_oficial = f"RLT.LAU-{num_raw}.26-00" if num_raw else "RLT.LAU-XXX.26-00"

            # Formata data
            data_raw     = data_emissao.strip()
            data_oficial = data_raw
            if '/' in data_raw:
                p = data_raw.split('/')
                if len(p) >= 2:
                    try:
                        dia = f"{int(p[0]):02d}"
                        mes = f"{int(p[1]):02d}"
                        ano = p[2] if len(p) == 3 else "2026"
                        data_oficial = f"{dia}/{mes}/{ano}"
                    except ValueError:
                        pass

            # Cabeçalho
            plan["T4"]  = rlt_oficial
            plan["F5"]  = cliente
            plan["F6"]  = obra
            plan["F7"]  = att
            plan["F8"]  = endereco
            plan["AQ7"] = data_oficial

            # Header
            plan.oddHeader.right.text   = f'&"Calibri,Bold"&11{rlt_oficial}'
            plan.oddFooter.right.text   = "&P / &N"
            plan.firstFooter.right.text = "&P / &N"

            # Bigorna
            colunas_golpes = ["F","G","H","I","J","K","L","M","N","O"]
            for i, val in enumerate(bigorna_vals):
                if val:
                    try:
                        plan[f"{colunas_golpes[i]}22"] = float(val.replace(',','.'))
                    except ValueError:
                        pass
            plan["T21"]  = media_big
            plan["AQ21"] = coef_big

            # Amostras
            borda_grossa = Side(style="medium")
            linha_atual  = 41
            for amostra in st.session_state.amostras:
                plan[f"B{linha_atual}"] = amostra["Amostra"]
                plan[f"O{linha_atual}"] = amostra["Posição"]
                for col, key in [("X","I.E. Médio"),("AB","I.E. Efetivo"),("AI","Resist. Estimada (MPa)")]:
                    val = amostra[key]
                    if val != "-":
                        try:
                            plan[f"{col}{linha_atual}"] = float(val.replace(',','.'))
                        except (ValueError, AttributeError):
                            plan[f"{col}{linha_atual}"] = "-"
                    else:
                        plan[f"{col}{linha_atual}"] = "-"
                plan[f"AS{linha_atual}"] = amostra["Dispersão"]
                linha_atual += 2

            # Bordas externas
            for row_idx in range(39, 164):
                cell = plan.cell(row=row_idx, column=45)
                b = copy(cell.border) if cell.border else Border()
                b.right = borda_grossa
                cell.border = b
            for col_idx in range(1, 46):
                cell = plan.cell(row=163, column=col_idx)
                b = copy(cell.border) if cell.border else Border()
                b.bottom = borda_grossa
                cell.border = b

            # Notas e responsável
            plan["B162"]  = notas
            plan["AB161"] = f"{resp_nome}\nCREA: {resp_crea}"
            plan["AB161"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="bottom")

            # Assinatura
            if assinatura is not None:
                try:
                    img        = xlImage(io.BytesIO(assinatura.read()))
                    img.width  = 160
                    img.height = 55
                    plan.add_image(img, "AJ162")
                except Exception:
                    pass

            # Salva na memória
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()

            st.download_button(
                label=f"⬇️ Baixar {rlt_oficial}.xlsx",
                data=excel_bytes,
                file_name=f"{rlt_oficial}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

        except FileNotFoundError:
            st.error("❌ Arquivo 'Modelo_esclerometria.xlsx' não encontrado no repositório.")
        except Exception as e:
            st.error(f"❌ Erro ao gerar o relatório: {e}")
